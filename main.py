from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from openpyxl.styles.colors import Color
import os
import sys

# Preparing script before we convert it to executable
# Use script directory when running as Python script, executable directory when compiled
if getattr(sys, 'frozen', False):
    application_path = os.path.dirname(sys.executable)
else:
    application_path = os.path.dirname(os.path.abspath(__file__))

# Putting together #2, #3, and #4 (input: raw_data.xlsx + month , output: Report with barchart, formulas and format)
month = input('Introduce month: ')

# Read workbook and select sheet
input_path = os.path.join(application_path, 'raw_data.xlsx')
wb = load_workbook(input_path)
sheet = wb.active  # Use the active sheet instead of hardcoded 'Report'

# Active rows and columns
min_column = wb.active.min_column
max_column = wb.active.max_column
min_row = wb.active.min_row
max_row = wb.active.max_row

# Instantiate a barchart
barchart = BarChart()

# Locate data and categories
data = Reference(sheet,
                 min_col=min_column+1,
                 max_col=max_column,
                 min_row=min_row,
                 max_row=max_row)  # including headers
categories = Reference(sheet,
                       min_col=min_column,
                       max_col=min_column,
                       min_row=min_row+1,
                       max_row=max_row)  # not including headers

# Adding data and categories
barchart.add_data(data, titles_from_data=True)
barchart.set_categories(categories)

# Make chart
sheet.add_chart(barchart, "B12")
barchart.title = 'Sales by Product line'
barchart.style = 5  # choose the chart style

# Write multiple formulas with a for loop
for i in range(min_column+1, max_column+1):  # (B, G+1)
    letter = get_column_letter(i)
    sheet[f'{letter}{max_row + 1}'] = f'=SUM({letter}{min_row + 1}:{letter}{max_row})'
    sheet[f'{letter}{max_row + 1}'].style = 'Currency'

# Add format
sheet['A1'] = 'Sales Report'
sheet['A2'] = month
sheet['A1'].font = Font('Arial', bold=True, size=20, color=Color(rgb='85B50B'))
sheet['A2'].font = Font('Arial', size=10)

# Create report directory if it doesn't exist
report_dir = os.path.join(application_path, 'report')
os.makedirs(report_dir, exist_ok=True)

output_path = os.path.join(report_dir, f'report_{month}.xlsx')
wb.save(output_path)