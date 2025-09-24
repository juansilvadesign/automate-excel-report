# Adds SUM Formulas Only
# Purpose: Takes raw data and adds SUM formulas at the bottom of each column

# Reads from raw_data.xlsx
# Adds =SUM() formulas below the last row of data
# Applies 'Currency' formatting to the totals
# Saves as report.xlsx

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import os

wb = load_workbook('raw_data.xlsx')
sheet = wb.active

min_column = wb.active.min_column
max_column = wb.active.max_column
min_row = wb.active.min_row
max_row = wb.active.max_row

#you can do it manually like this:
'''
sheet['B8'] = '=SUM(B6:B7)
sheet['B8'].style = 'Currency'
'''

for i in range(min_column+1, max_column+1):
    print(i)
    letter = get_column_letter(i)
    sheet[f'{letter}{max_row+1}'] = f'=SUM({letter}{min_row+1}:{letter}{max_row})'
    sheet[f'{letter}{max_row+1}'].style = 'Currency'

# Create report directory if it doesn't exist
os.makedirs('report', exist_ok=True)
wb.save('report/report.xlsx')
